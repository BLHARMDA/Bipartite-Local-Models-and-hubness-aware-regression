# coding=utf-8
# # # # # # # # # # # # # # # #
# HLM
# language:Python2.7
# env:Anaconda2
# # # # # # # # # # # # # # # #

import numpy as np
import scipy as sp
import scipy.sparse
import openpyxl as xlwt
import xlrd
import random
import datetime
import copy
import numba
import sys,os
reload(sys) 
sys.setdefaultencoding('gbk') 

#################################################

# # # # # data path # # # # #
full_path=os.path.realpath(__file__)
eop=full_path.rfind(__file__)
folder_path=full_path[0:eop]+u'数据'

# # # # # global # # # # #

M_num=495 #miRNA num
D_num=383 #disease num
MD_num=5430 #association num

# md
A=[['0'],['0']] 
m_d=np.zeros(1) #m-d
m_d_backup=np.zeros(1) #backup for A
# ss
ss=np.zeros(1) #ss
ss_w=np.zeros(1) #ss weight
# fs
fs=np.zeros(1) #fs
fs_w=np.zeros(1) #fs weight

y_local=np.zeros(1)
y_global=np.zeros(1)

######################################################################

# md
test_path_md=folder_path+u'/1.miRNA-disease关联数据/miRNA-disease关联数字编号.xlsx'
# ss
test_path_ss_1=folder_path+u'/2.disease semantic similarity 1/疾病语义类似性矩阵1.txt'
test_path_ss_2=folder_path+u'/3.disease semantic similarity 2/疾病语义类似性矩阵2.txt'
test_path_ss_w=folder_path+u'/2.disease semantic similarity 1/疾病语义类似性加权矩阵1.txt'
# fs
test_path_fs=folder_path+u'/4.miNA功能类似性矩阵/miRNA功能类似性矩阵.txt'
test_path_fs_w=folder_path+u'/4.miNA功能类似性矩阵/miRNA功能类似性加权矩阵.txt'

######################################################################

# # # # # # # # I n i t # # # # # # # #
# Data Loader, Preprocess
# Output:m_m,d-d,d_d,m_d,M_num,D_num,M,D
# Packaging:__start__()
# # # # # # # # # # # # # # # # # # # # # # #

# # # # # Packaging # # # # #

def __start__(path=folder_path,nm=M_num,nd=D_num)
    global folder_path,M_num,D_num
    folder_path=path
    M_num=nm
    D_num=nd
    load_md(test_path_md)
    load_ss(test_path_ss_1,test_path_ss_2,test_path_ss_w)
    load_fs(test_path_fs,test_path_fs_w)

# # # # # Data Loader # # # # #

def load_md(path_md):
    
    global M_num,D_num,MD_num
    global A,m_d,m_d_backup
    
    # load data from excel
    md_table=xlrd.open_workbook(path_md)
    md_sheet=md_table.sheet_by_index(0)
    MD_num=md_sheet.nrows
    
    m_d=np.zeros([M_num,D_num]) #M x D
    for i in range(MD_num):
        row_index=int(md_sheet.cell_value(rowx=i,colx=0))-1
        col_index=int(md_sheet.cell_value(rowx=i,colx=1))-1
        A[0].append(int(row_index))
        A[1].append(int(col_index))
        m_d[row_index,col_index]=1
    m_d_backup=copy.copy(m_d)
    # remove initial data
    del A[0][0],A[1][0]

def load_ss(path_ss_1,path_ss_2,path_ss_w):
    
    global ss,ss_w
    
    # load data from txt
    ss_1=np.loadtxt(path_ss_1)
    ss_2=np.loadtxt(path_ss_2)
    ss_w=np.loadtxt(path_ss_w)
    
    ss=(ss_1+ss_2)/2

def load_fs(path_fs,path_fs_w):
    
    global fs,fs_w
    
    fs=np.loadtxt(path_fs)
    fs_w=np.loadtxt(path_fs_w)

######################################################################

# # # numba # # # 

@numba.jit
def jitsum(x): #Summing up by row
    [m,n]=x.shape
    s=np.zeros(m)
    for i in range(int(m)):
        for j in range(int(n)):
            s[i]+=x[i,j]
    return s
@numba.jit
def jitsumt(x): #Summing up by column
    [m,n]=x.shape
    s=np.zeros(n)
    for i in range(int(m)):
        for j in range(int(n)):
            s[j]+=x[i,j]
    return s
@numba.jit
def jitsumall(x):
    [m,n]=x.shape
    s=0
    for i in range(int(m)):
        for j in range(int(n)):
            s+=x[i,j]
    return s

# # # # # Preprocess # # # # #

@numba.jit
def __init__():
    
    global M_num,D_num
    global ss,ss_w,fs,fs_w,m_d
    
    ''' Gaussian interaction Profile '''
    #Gamma
    Gamma_d_s=1
    Gamma_m_s=1
    md_f=jitsumall(m_d*m_d)
    Gamma_d=Gamma_d_s/(md_f/D_num)
    Gamma_m=Gamma_m_s/(md_f/M_num)
    
    #KD
    IP_d=np.tile(m_d.T,[1,D_num])-np.resize(m_d.T,[1,M_num*D_num]) #DxMD
    IP_d=np.resize(IP_d*IP_d,[D_num*D_num,M_num])
    gs_d=np.exp(-Gamma_d*np.resize(jitsum(IP_d),[D_num,D_num]))
    #KM
    IP_m=np.tile(m_d,[1,M_num])-np.resize(m_d,[1,M_num*D_num]) #MxMD
    IP_m=np.resize(IP_m*IP_m,[M_num*M_num,D_num])
    gs_m=np.exp(-Gamma_m*np.resize(jitsum(IP_m),[M_num,M_num]))

    m_m=fs*fs_w+(1-fs_w)*gs_m
    d_d=ss*ss_w+(1-ss_w)*gs_d
    
    #Jaccob
    m_j=np.ones([M_num,M_num])
    d_j=np.ones([D_num,D_num])
    
    Mi=np.dot(m_d,m_d.T) #Intersect
    Di=np.dot(m_d.T,m_d)
    Mu_t=np.tile(m_d,[1,M_num])+np.resize(m_d,[1,M_num*D_num]) #Union
    Mu_t=np.resize(Mu_t,[M_num*M_num,D_num])
    Mu=np.resize(jitsum(Mu_t),[M_num,M_num])-Mi
    Du_t=np.tile(m_d.T,[1,D_num])+np.resize(m_d.T,[1,M_num*D_num])
    Du_t=np.resize(Du_t,[D_num*D_num,M_num])
    Du=np.resize(jitsum(Du_t),[D_num,D_num])-Di
    m_j=Mi/(Mu+(Mu==0))
    d_j=Di/(Du+(Du==0))
    
    # Integrate
    M=np.column_stack((m_m,m_j))
    D=np.column_stack((d_d,d_j))

    return M,D
########################################################################

# # # # # H L M # # # # #

# Ranking by Euclidean distance, 1 for the instances ranked less than k
@numba.jit
def kNNmat(rankdata,k): #kNN
    num=int(rankdata.shape[0])
    rank=np.zeros([num,num])
    for i in range(num):
        rki=np.argsort(rankdata[i])
        ranki=np.zeros(num)
        for j in range(num):
            t=rki[j]
            ranki[t]=j
        for j in range(num):
            rank[i,j]=1 if ranki[j]<k else 0
    return rank

@numba.jit
def MatEu(mat): #Euclidean distance mat
    num=int(mat.shape[0])
    mt=np.tile(mat,[1,num])-np.resize(mat,[1,num*2*num]) #a-b
    mt=np.resize(mt*mt,[num*num,2*num])
    mt=np.resize(jitsum(mt),[num,num])
    return mt

@numba.jit
def HLM(F_M,F_D,k,M,D):
    
    global M_num,D_num,m_d
    y=m_d
    yc_w=0
    yc_k=0
    IP=np.tile(jitsumt(m_d)==0,[M_num,1])+np.tile(jitsum(m_d)==0,[D_num,1]).T!=0
    Mx=np.tile(m_d,[2,1]) #2MxD
    Dx=np.tile(m_d,[1,2]) #Mx2D
    
    j_m=F_M
    j_d=F_D
    M_rd=M
    D_rd=D
    
    #cal Euclidean distance
    M_eu=MatEu(M_rd)
    D_eu=MatEu(D_rd)
    #ranking
    knn_m=kNNmat(M_eu,k) #MxM
    knn_d=kNNmat(D_eu,k) #DxD
    
    ''' ECkNN '''
    yc_m=np.dot(knn_m.T,y) #RkNN x m_d MxD
    yc_d=np.dot(y,knn_d) #MxD
    yc_m_2=np.tile(jitsumt(knn_m),[D_num,1]).T
    yc_d_2=np.tile(jitsumt(knn_d),[M_num,1])
    
    #yc_d yc_m
    yc_d+=(yc_d_2==0)*y
    yc_d_2+=(yc_d_2==0)
    yc_d/=yc_d_2
    
    yc_m+=(yc_m_2==0)*y
    yc_m_2+=(yc_m_2==0)
    yc_m/=yc_m_2
    
    #y_d y_m
    x_m=np.resize(yc_m.T,[1,D_num*M_num])*np.tile(knn_m,[1,D_num]) #MxMD
    x_d=np.resize(yc_d,[1,M_num*D_num])*np.tile(knn_d,[1,M_num]) #DxDM
    sum_m=jitsum(np.resize(x_m,[M_num*D_num,M_num])) #1xMD
    sum_d=jitsum(np.resize(x_d,[D_num*M_num,D_num]))
    y_m=np.resize(sum_m,[M_num,D_num])
    y_d=np.resize(sum_d,[D_num,M_num]).T
    
    yc_k=(y_m+y_d)/(2.0*k)
    
    ''' weighted '''
    # yc_1 yc_2
    yc_1_b=jitsum(M_rd) #Mx1
    yc_2_b=jitsum(D_rd) #Dx1
    yc_1=np.dot(M_rd,Mx).T/(yc_1_b+(yc_1_b==0)) #DxM
    yc_2=np.dot(Dx,D_rd.T)/(yc_2_b+(yc_2_b==0)) #MxD

    yc_w=(yc_1.T+yc_2)/2.0

    ''' Aggregate '''
    yc=IP*yc_w+(1-IP)*yc_k
    
    return yc

# # # # # # # # # # # # cross validation # # # # # # # # # # # # # #

@numba.jit
def rank(p,req_point):
    p_r=sorted(p,reverse=True)
    q=0 #record repetitions
    rank=1
    for i in range(len(p)):
        if p_r[i]==req_point:
            rank=i+1
            q+=1
    return (2*rank+1-q)/2.0

# # # # # # L O O C V # # # # # #

def LOOCV(F_M,F_D,k,N):
    
    global M_num,D_num,MD_num
    global m_d,A,y_local,y_global
    y_local=[]
    y_global=[]
    
    time=0
    for i in range(Div_start,Div_end):
        t1=datetime.datetime.now()
        m_d=copy.copy(m_d_backup) #load m-d
        m_d[int(A[0][i]),int(A[1][i])]=0 #remove ith instance in m_d
        [M,D]=__init__()
        yc=HLM(F_M,F_D,k,N,M,D)
        score=yc[int(A[0][i]),int(A[1][i])]
        
        # Local
        yc_p=[]
        for j in range(M_num):
            if m_d[j,int(A[1][i])]==0:
                yc_p.append(yc[j,int(A[1][i])])
        y_local.append(rank(yc_p,score))
        
        # Global
        yc_p=[]
        for j in range(M_num):
            for k in range(D_num):
                if m_d[j,k]==0:
                    yc_p.append(yc[j,k])
        y_global.append(rank(yc_p,score))
        
        t2=datetime.datetime.now()
        print 'Iterration:',i+1,' Time:',t2-t1,' Local:',y_local[time],' Global:',y_global[time],'\n'
        time+=1

########### Output #############

def __end__():
    
    local_table = xlwt.Workbook()
    global_table = xlwt.Workbook()
    local_sheet=local_table.active
    global_sheet=global_table.active
    local_sheet.append(y_local)
    global_sheet.append(y_global)


    save_path=folder_path+'/merge'
    # Save
    path1=save_path+'/local'+str(dist_id)+'.xlsx'
    path2=save_path+'/global'+str(dist_id)+'.xlsx'
    local_table.save(path1)
    global_table.save(path2)

# # # # # tools # # # # #

def test_speed(F_M=495*2,F_D=383*2,k=100,N=1):
    t1=datetime.datetime.now()
    # init
    global M_num,D_num,MD_num
    global m_d,A
    y_local=np.zeros(MD_num)
    y_global=np.zeros(MD_num)
    i=random.sample(range(MD_num),1)[0]
    m_d=copy.copy(m_d_backup)
    
    m_d[int(A[0][i]),int(A[1][i])]=0
    [M,D]=__init__()
    yc=HLM(F_M,F_D,k,N,M,D)
    score=yc[int(A[0][i]),int(A[1][i])]
    
    # Local
    yc_p=[score]
    for j in range(M_num):
        if m_d[j,int(A[1][i])]==0:
            yc_p.append(yc[j,int(A[1][i])])
    y_local[i]=rank(yc_p,score)

    # Global
    yc_p=[score]
    for j in range(M_num):
        for k in range(D_num):
            if m_d[j,k]==0:
                yc_p.append(yc[j,k])
    y_global[i]=rank(yc_p,score)

    t2=datetime.datetime.now()
    return t2-t1

''' simple distribute tool '''
dist_id=0 #id for this console, start from 0
Div_num=2000 #working load for each console

Div_start=0
Div_end=0

def dist_set(idi):
    global dist_num,dist_id,Div_num,Div_start,Div_end
    dist_id=idi
    Div_start=Div_num*dist_id
    if MD_num-Div_start<Div_num:
        Div_end=MD_num
    else:
        Div_end=Div_start+Div_num

# Merge output
def dist_merge():
    global Div_num,Div_start,Div_end,local_data,global_data
    q=MD_num/Div_num+1
    id_now=0
    path=folder_path+'/merge'
    save_path=folder_path
    local_data=[]
    global_data=[]
    while id_now<q:
        path_local=path+'/local'+str(id_now)+'.xlsx'
        path_global=path+'/global'+str(id_now)+'.xlsx'
        # read from excel
        local_table=xlrd.open_workbook(path_local)
        global_table=xlrd.open_workbook(path_global)
        local_sheet=local_table.sheet_by_index(0)
        global_sheet=global_table.sheet_by_index(0)
        
        Cir=MD_num%Div_num if id_now==q-1 else Div_num
        for i in range(Cir):
            local_data.append(local_sheet.cell_value(rowx=0,colx=i))
            global_data.append(global_sheet.cell_value(rowx=0,colx=i))
        
        id_now+=1
    
    local_table = xlwt.Workbook()
    global_table = xlwt.Workbook()
    local_sheet=local_table.active
    global_sheet=global_table.active
    local_sheet.append(local_data)
    global_sheet.append(global_data) 

    # save
    path1=save_path+'/local.xlsx'
    path2=save_path+'/global.xlsx'
    local_table.save(path1)
    global_table.save(path2)
    print 'Merge finished!\n'

'''start func for loocv '''
def __run__(F_M,F_D,k,N,id):
    path=folder_path+'/merge'
    if not os.path.exists(path):
        os.mkdir(path)

    print 'Loading data...\n'
    __start__(folder_path,495,383)
    dist_set(id)
    print 'Data loaded.\n'
    print 'Your Districted Id:',dist_id,' Max id:',5430/Div_num,' Working amount:',Div_end-Div_start,'\n'
    print 'Compiling...\n'
    test_speed(F_M,F_D,k,N)
    print 'Calculating running time...\n'
    time=test_speed(F_M,F_D,k,N)
    print 'Data loaded. Runing time:',time*(Div_end-Div_start)/3600,'hours\n\nLOOCV start.\n'
    LOOCV(F_M,F_D,k,N)
    print 'LOOCV finished, start to output. Output format: xlsx\n'
    __end__()
    print 'File output finished. Over.\n'

def single_run(F_M,F_D,k,N): #not use distribute
    global Div_start,Div_end
    
    Div_start=0
    Div_end=5430
    print 'Loading data...\n'
    __start__(folder_path,495,383)
    print 'Compiling...\n'
    test_speed(F_M,F_D,k,N)
    print 'Data loaded, calculating running time...'
    time=test_speed(F_M,F_D,k,N)
    print 'Runing time: ',time*5430/3600,' hours.\nLOOCV start\n'
    LOOCV(F_M,F_D,k,N)
    print 'LOOCV finished, start to output. Output format: xlsx\n'

    ''' save '''
    local_table = xlwt.Workbook()
    global_table = xlwt.Workbook()
    local_sheet=local_table.active
    global_sheet=global_table.active
    local_sheet.append(y_local)
    global_sheet.append(y_global)

    path1=folder_path+'/local.xlsx'
    path2=folder_path+'/global.xlsx'
    local_table.save(path1)
    global_table.save(path2)
    print 'File output finished. Over.\n'

''' M a i n '''
#my_id=int(sys.argv[1])
#__run__(495*2,383*2,100,1,my_id)

''' merge '''
#dist_merge()

''' single console '''
#single_run(495*2,383*2,100,1)

#############################################################################################

# # # # # # k - f o l d # # # # # #
dist_quant_for_fold=50 #setting for distribute

@numba.jit
def total_rank(p):
    p_a=np.array(p)
    p_r=sorted(p)
    rk=np.argsort(p_a)
    num=len(p)
    rkf=np.zeros(num)
    for i in range(num):
        t=rk[i]
        rkf[t]=i
    q=0
    save=0
    for i in range(1,num):
        if(p_r[i]==p_r[i-1]):
            q+=1
            save=p_r[i]
        else:
            if q!=0:
                indexp=p.index(save)
                st_rank=rkf[indexp]
                rank=(2*st_rank+q)/2.0
                for j in range(num):
                    if save==p[j]:
                        rkf[j]=rank
                q=0
    return num-rkf

def fold(F_M,F_D,k,N,did=-1):
    
    global M_num,D_num,MD_num
    global m_d,A,y_global

    table = xlwt.Workbook()
    sheet=table.active
    # each console run Num times
    Num=100 if did==-1 else dist_quant_for_fold
    fold_num=MD_num/5
    for i in range(Num):
        print 'Iterration ',i,'\n'
        rand_A=np.arange(MD_num)
        random.shuffle(rand_A)
        y_global=np.zeros(MD_num)
        for j in range(5):
            t1=datetime.datetime.now()
            m_d=copy.copy(m_d_backup)
            for chs in range(fold_num*j,fold_num*(j+1)):
                index=rand_A[chs]
                m_d[int(A[0][index]),int(A[1][index])]=0
            [M,D]=__init__()
            yc=HLM(F_M,F_D,k,N,M,D)

            yc_p=[]
            for r in range(M_num):
                for k in range(D_num):
                    if m_d[r,k]==0:
                        yc_p.append(yc[r,k])
            rank_p=total_rank(yc_p)

            for chs in range(fold_num*j,fold_num*(j+1)):
                index=rand_A[chs]
                score=yc[int(A[0][index]),int(A[1][index])]
                index_i=yc_p.index(score)
                y_global[index]=rank_p[index_i]
            t2=datetime.datetime.now()
            print 'fold:',j,' Time:',t2-t1,'\n'
        sheet.append(y_global.tolist())
    save_path=folder_path+'/merge'
    #save,did=-1 for single console
    path=folder_path+'/fold.xlsx' if did==-1 else save_path+'/fold'+str(did)+'.xlsx' 
    table.save(path)

def run_for_fold(F_M,F_D,k,N,did):
    path=folder_path+'/merge'
    if not os.path.exists(path):
        os.mkdir(path)
        
    print 'Loading data...\n'
    __start__(folder_path,495,383)
    print 'Data loaded.\n'
    print 'Your Districted Id:',did,' Max id:',100/dist_quant_for_fold-1,' Working amount:',5*dist_quant_for_fold,'\n'
    print 'Compiling...\n'
    t=np.random.rand(10)
    total_rank(t)
    test_speed(F_M,F_D,k,N)
    print 'Calculating running time...\n'
    time=64.0/3600
    run_time=time*100 if did==-1 else time*dist_quant_for_fold
    print 'Data loaded. Runing time:',run_time,'hours\n\nfold start.\n'
    fold(F_M,F_D,k,N,did)
    print 'fold finished, start to output. Output format: xlsx\n'
    print 'File output finished. Over.\n'

# merge
def merge_for_fold():
    global MD_num,global_data
    q=100/dist_quant_for_fold
    id_now=0
    path=folder_path+'/merge'
    save_path=folder_path
    table = xlwt.Workbook()
    sheet=table.active
    while id_now<q:
        path_global=path+'/fold'+str(id_now)+'.xlsx'
        global_table=xlrd.open_workbook(path_global)
        global_sheet=global_table.sheet_by_index(0)
        for rowx in range(dist_quant_for_fold):
            global_data=[]
            for i in range(MD_num):
                global_data.append(global_sheet.cell_value(rowx=rowx,colx=i))
            sheet.append(global_data) 
        id_now+=1
    path=save_path+'/kfold.xlsx'
    table.save(path)
    print 'Merge finished!\n'

''' run '''
#my_id=int(sys.argv[1]) #set id,-1 for not use distribute
#run_for_fold(495*2,383*2,100,1,my_id)

''' merge '''
#merge_for_fold()
